"""
Multi-format Document Loaders

Dispatcher that loads PDF, DOCX, XLSX, Email (.eml/.msg), and TXT files
into LangChain Document objects with appropriate metadata.
"""
import email
from email import policy
from pathlib import Path

from langchain_core.documents import Document
from langchain_community.document_loaders import PyPDFLoader

# Supported file extensions
SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".eml", ".msg", ".txt"}


def load_document(file_path: str | Path) -> list[Document]:
    """
    Load any supported file and return raw page/section documents.

    Args:
        file_path: Path to the document file.

    Returns:
        List of Document objects with metadata.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file format is not supported.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    ext = file_path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file format: {ext}. "
            f"Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )

    print(f"📄 Loading: {file_path.name}")

    if ext == ".pdf":
        docs = _load_pdf(file_path)
    elif ext == ".docx":
        docs = _load_docx(file_path)
    elif ext == ".xlsx":
        docs = _load_xlsx(file_path)
    elif ext == ".eml":
        docs = _load_eml(file_path)
    elif ext == ".msg":
        docs = _load_msg(file_path)
    elif ext == ".txt":
        docs = _load_txt(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext}")

    print(f"   {len(docs)} page(s)/section(s) loaded.")
    return docs


def _load_pdf(file_path: Path) -> list[Document]:
    """Load PDF using PyPDFLoader."""
    loader = PyPDFLoader(str(file_path))
    return loader.load()


def _load_docx(file_path: Path) -> list[Document]:
    """Load DOCX using docx2txt."""
    import docx2txt

    text = docx2txt.process(str(file_path))
    if not text or not text.strip():
        return []

    return [Document(
        page_content=text,
        metadata={"source": str(file_path)},
    )]


def _load_xlsx(file_path: Path) -> list[Document]:
    """Load XLSX using openpyxl -- one Document per sheet as a markdown table."""
    from openpyxl import load_workbook

    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    docs = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find the first non-empty row as header
        header_idx = 0
        for i, row in enumerate(rows):
            if any(cell is not None for cell in row):
                header_idx = i
                break

        headers = [str(cell) if cell is not None else "" for cell in rows[header_idx]]
        data_rows = rows[header_idx + 1:]

        if not data_rows:
            continue

        # Build markdown table
        lines = []
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join("---" for _ in headers) + " |")

        for row in data_rows:
            if not any(cell is not None for cell in row):
                continue  # Skip empty rows
            cells = [str(cell) if cell is not None else "" for cell in row]
            lines.append("| " + " | ".join(cells) + " |")

        if len(lines) > 2:  # Has at least header + separator + one data row
            docs.append(Document(
                page_content="\n".join(lines),
                metadata={
                    "source": str(file_path),
                    "sheet_name": sheet_name,
                },
            ))

    wb.close()
    return docs


def _load_eml(file_path: Path) -> list[Document]:
    """Load .eml email file using Python's email module."""
    with open(file_path, "rb") as f:
        msg = email.message_from_binary_file(f, policy=policy.default)

    subject = msg.get("Subject", "")
    from_addr = msg.get("From", "")
    to_addr = msg.get("To", "")
    date = msg.get("Date", "")

    # Extract body text
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            if content_type == "text/plain":
                body = part.get_content()
                break
            elif content_type == "text/html" and not body:
                body = part.get_content()
    else:
        body = msg.get_content()

    if not body or not str(body).strip():
        return []

    # Build a readable header block
    header_block = ""
    if subject:
        header_block += f"Subject: {subject}\n"
    if from_addr:
        header_block += f"From: {from_addr}\n"
    if to_addr:
        header_block += f"To: {to_addr}\n"
    if date:
        header_block += f"Date: {date}\n"
    if header_block:
        header_block += "\n"

    return [Document(
        page_content=header_block + str(body),
        metadata={
            "source": str(file_path),
            "email_subject": subject,
            "email_from": from_addr,
            "email_to": to_addr,
            "email_date": date,
        },
    )]


def _load_msg(file_path: Path) -> list[Document]:
    """Load Outlook .msg file using extract-msg."""
    import extract_msg

    msg = extract_msg.Message(str(file_path))
    subject = msg.subject or ""
    sender = msg.sender or ""
    to = msg.to or ""
    date = str(msg.date) if msg.date else ""
    body = msg.body or ""

    msg.close()

    if not body.strip():
        return []

    header_block = ""
    if subject:
        header_block += f"Subject: {subject}\n"
    if sender:
        header_block += f"From: {sender}\n"
    if to:
        header_block += f"To: {to}\n"
    if date:
        header_block += f"Date: {date}\n"
    if header_block:
        header_block += "\n"

    return [Document(
        page_content=header_block + body,
        metadata={
            "source": str(file_path),
            "email_subject": subject,
            "email_from": sender,
            "email_to": to,
            "email_date": date,
        },
    )]


def _load_txt(file_path: Path) -> list[Document]:
    """Load plain text file."""
    text = file_path.read_text(encoding="utf-8", errors="replace")
    if not text.strip():
        return []

    return [Document(
        page_content=text,
        metadata={"source": str(file_path)},
    )]
